import openpyxl as xl
import pandas as pd



# 엑셀 파일 초기화 후 재입력
path = r''

wb = xl.load_workbook()
second = wb.worksheets[1]

for row in second.iter_rows():
    for cell in row:
        cell.value = None

temp = [2000]
df = pd.DataFrame(temp)

for idx, row in df.iterrows():
    for col_idx, value in enumerate(row, start=1):
        cell = second.cell(row=idx + 1, column=col_idx, value=value)

wb.save(filename=r'C:\Users\82102\Desktop\test.xlsb')



def check_size_after_merge(func):
    '''
    병합 전후 사이즈 검증 데코레이터
    '''
    def wrapper(df, *args, **kwargs):
        initial = df.shape
        result = func(*args, **kwargs)
        final = result.shape
        
        if initial == final:
            return print('병합 오류')
        else:
            return result
    return wrapper