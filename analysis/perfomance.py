from analysis.monthly_close import PaidProductCount
from config import DATA_PATH
from data.columns import member_main_column, program_cate_column
from datetime import datetime
import numpy as np
import pandas as pd
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
import os


class MainIndex:
    '''
    누적 가입자, active 유저 등 기초 정보 추출
    
    그룹화 전의 데이터
    '''
    member_table_path = DATA_PATH['member_list']
    cate_table_path = DATA_PATH['program_category']
    
    def __init__(self):
        self.table = self.merge_tables()
    
    def create_tables(self):
        '''
        가입자 테이블과 상품정보 테이블 딕셔너리로 반환
        '''
        tables = {}
        
        member = pd.read_parquet(self.member_table_path)
        col_member = member_main_column()
        member = member[col_member]
        tables['member'] = member
        print('가입자 테이블 생성')
    
        cate = pd.read_parquet(self.cate_table_path)
        col_cate = program_cate_column()
        cate = cate[col_cate]
        tables['category'] = cate
        print('상품정보 테이블 생성')
        return tables

    def merge_tables(self):
        '''
        딕셔너리의 자료를 병합
        '''
        tables_dic = self.create_tables()
        mg = pd.merge(tables_dic['member'], tables_dic['category'], how='left', on='상품정보')        
        print('가입자/상품정보 테이블 병합 중...')
        print(f'병합 전 가입자 테이블 크기 : {tables_dic["member"].shape[0]}')
        print(f'병합 후 가입자 테이블 크기 : {mg.shape[0]}')
        
        if tables_dic["member"].shape[0] != mg.shape[0]:
            mg = mg.drop_duplicates()
            print('병합 후 중복 데이터 제거')
        
        print('병합 이상 없음 --> 작업 완료')
        
        mg = self.create_additional_columns(mg)
        return mg
    
    def create_additional_columns(self, df_join):
        '''
        그룹화를 위한 추가 칼럼 생성
        
        보험가입월, 보험해지월, active 유저
        '''
        print('기간 정보 추가 중...')
        # 가입기간 추가
        df_join['보험해지일'] = pd.to_datetime(df_join['보험해지일'])
        df_join['보험가입일'] = pd.to_datetime(df_join['보험가입일'])
        # df_join = df_join[df_join['보험가입일'] <= '2023-09-15']
        print('보험유지기간 계산 중...')
        df_join['보험유지기간'] = (df_join['보험해지일'].dt.year - df_join['보험가입일'].dt.year) * 12 + (df_join['보험해지일'].dt.month - df_join['보험가입일'].dt.month)
        
        # 유지기간 분류 추가
        print('유지기간 분류 중... : 12개월 이하, 14개월 이하 등...')
        # df_join['유지분류'] = df_join['보험유지기간'].apply(self.check_hold)
        holding_conditions = [
            df_join['보험유지기간'] <= 12,
            df_join['보험유지기간'] <= 14,
            df_join['보험유지기간'] <= 18,
            df_join['보험유지기간'] <= 24,
            df_join['보험유지기간'] >= 25
        ]
        values = [12, 14, 18, 24, 25]
        df_join['유지분류'] = np.select(holding_conditions, values, '유지중')
        
        
        # 월별 가입 추가
        df_join['보험가입월'] = df_join['보험가입일'].dt.strftime('%Y-%m')
        df_join['보험해지월'] = df_join['보험해지일'].dt.strftime('%Y-%m')
        print('기간 정보 추가 : 보험유지기간, 보험가입월, 보험해지월')

        # active 유저 추가
        active_condition = (df_join['보험상태'].isin(['가입', '해지예약', '미납']))
        df_join['active'] = np.where(active_condition, True, False)
        # df_join['active'] = df_join['보험상태'].apply(self.active_user)
        print('ACTIVE 유저 판정 정보 생성')
        
        # 미납 분류
        unpaid_condition = (df_join['보험상태'] == '미납')
        df_join['미납여부'] = np.where(unpaid_condition, True, False)
        
        return df_join
    
    def __check_hold(self, row):
        '''
        유지기간 분류 도우미 함수
        
        정수 타입 반환
        '''
        if row <= 12:
            return 12
        elif row <= 14:
            return 14
        elif row <= 18:
            return 18
        elif row <= 24:
            return 24
        elif row >= 25:
            return 25
        else:
            return "유지중"
    
    def __active_user(self, row):
        '''
        active 유저 구분 도우미 함수
        '''
        if (row == '가입') or (row == '해지예약'):
            return True
        else:
            return False


class MonthlyJoin(MainIndex):
    '''
    MainIndex 클래스에서 추출한 데이터를 그룹화
    
    신규, 해지, 제품별, active 등
    '''

    user_group = [
        '보험가입월',
        '프로모션유무',
        'active',
        '미납여부'
    ]
    cancel_group = [
        '보험해지월',
        '프로모션유무',
        'active',
        '미납여부'
    ]
    product_group = [
        '보험가입월',
        '프로모션유무',
        'active',
        '보장타입',
        '제품시리즈',
        '미납여부'
    ]
    
    def __init__(self, save: bool):
        super().__init__()
        self.result_tables = self.result_table_dic()
        self.save_dataframe(save)

    def new_user(self):
        join = pd.DataFrame(self.table.groupby(self.user_group)['상품정보'].count()).reset_index()
        print('신규 고객 집계 완료')
        return join
    
    def cancel_user(self):
        cancel = pd.DataFrame(self.table.groupby(self.cancel_group)['상품정보'].count()).reset_index()
        print('해지 고객 집계 완료')
        return cancel
    
    def monthly_user_table(self):
        join = self.new_user()
        cancel = self.cancel_user()
        left_on = ['보험가입월', '프로모션유무', 'active', '미납여부']
        right_on = ['보험해지월', '프로모션유무', 'active', '미납여부']
        suffixes = ['가입', '해지']
        mg = pd.merge(join, cancel, left_on=left_on, right_on=right_on, how='left', suffixes=suffixes)
        mg['년'] = mg['보험가입월'].str[:4] + '년'
        mg['월'] = mg['보험가입월'].str[5:].astype(int).astype(str) + '월'
        print('신규/해지 고객 그룹화 테이블 생성')
        return mg
    
    def monthly_product_table(self):
        product = pd.DataFrame(self.table.groupby(self.product_group)['상품정보'].count()).reset_index()
        product['년'] = product['보험가입월'].str[:4] + '년'
        product['월'] = product['보험가입월'].str[5:].astype(int).astype(str) + '월'
        print('상품별 그룹화 테이블 생성')
        return product
    
    def holding_user_table(self):
        holding = pd.DataFrame(self.table.groupby(['프로모션유무', '제품군_2', '유지분류', 'active', '미납여부'])['상품정보'].count().reset_index())
        print('가입 유지기간 그룹화 테이블 생성')
        return holding
    
    def payment_number_table(self):
        close = PaidProductCount.ProcessAdditionalData(False)

        # 당월 결제 수량
        pay_total = close.dataframe.groupby(['제품군_2', '보장타입', '유무상'])['주문번호'].count().reset_index()
        pay_total.rename(columns={'주문번호': '총결제수량'}, inplace=True)
        # 당월 결제 수량(중복결제 제외)
        pay_not_dupl = close.dataframe.drop_duplicates('POLICY_ID').groupby(['제품군_2', '보장타입', '유무상'])['주문번호'].count().reset_index()
        pay_not_dupl.rename(columns={'주문번호': '중복제외수량'}, inplace=True)

        merge = pd.merge(pay_total, pay_not_dupl, how='left', on=['제품군_2', '보장타입', '유무상'])
        print('결제 건수 그룹화 테이블 생성')
        return merge
    
    def result_table_dic(self):
        dic = {}
        user = self.monthly_user_table()
        product = self.monthly_product_table()
        holding = self.holding_user_table()
        payment = self.payment_number_table()
        dic['user_pivot'] = user
        dic['product_pivot'] = product
        dic['holding_pivot'] = holding
        dic['payment_number'] = payment
        print('결과 테이블 딕셔너리 반환')
        return dic
    
    def save_dataframe(self, save: bool):
        if save:
            new_sht = [key for key in self.result_tables.keys()]
            print('엑셀 파일 생성 중...')
            wb = xl.Workbook()
            
            for sht in new_sht:
                print(f'{sht} 시트 생성 중...')
                ws = wb.create_sheet(sht)
                    
                print('데이터 입력 중 : [', end='')
                ratio = self.result_tables[sht].shape[0] * 0.01
                for num, row in enumerate(dataframe_to_rows(self.result_tables[sht], index=False, header=True)):
                    ws.append(row)
                    
                    if num >= ratio:
                        print('#', end='')
                        ratio += self.result_tables[sht].shape[0] * 0.01
                print(f']\n{sht} 시트 입력 완료')
            
            delete = wb['Sheet']
            wb.remove(delete)
            
            print('데이터 저장 중...')
            now = datetime.now().strftime('%Y%m%d')
            wb.save(fr'result\주요지표현황_{now}.xlsx')
            print('데이터 저장 완료!!!')


class ConversionRate:
    '''
    프로모션 유상 전환율
    
    Parameter
    ---
    start : '2023-08-01' 형식 / 마감 월에 맞춰 입력
    
    end : '2023-08-31' 형식 / 마감 월에 맞춰 입력
    '''
    def __init__(self, start: str, end: str=None, save: bool=True):
        self.date = start
        self.conversion_dic = self.__extract_data(start, end)
        self.__save_dataframe(save)
    
    def __extract_data(self, start: str, end: str):
        conversion_dic = {}
        pid = ConversionRate.FirstPayment(start, end).result
        exp = ConversionRate.ExpiredProduct(start, end).expiration_group
        conversion_dic['pid'] = pid
        conversion_dic['exp'] = exp
        return conversion_dic
    
    def __save_dataframe(self, save: bool):
        if save:
            new_sht = [key for key in self.conversion_dic.keys()]
            print('엑셀 파일 생성 중...')
            wb = xl.Workbook()
            
            for sht in new_sht:
                print(f'{sht} 시트 생성 중...')
                ws = wb.create_sheet(sht)
                    
                print('데이터 입력 중 : [', end='')
                ratio = len(self.conversion_dic[sht]) * 0.01
                for num, row in enumerate(dataframe_to_rows(self.conversion_dic[sht], index=False, header=True)):
                    ws.append(row)
                    
                    if num >= ratio:
                        print('#', end='')
                        ratio += len(self.conversion_dic[sht]) * 0.01
                print(f']\n{sht} 시트 입력 완료')
            
            delete = wb['Sheet']
            wb.remove(delete)
            
            print('데이터 저장 중...')
            now = datetime.now().strftime('%Y%m%d')
            wb.save(fr'result\프로모션전환율_{now}.xlsx')
            print('데이터 저장 완료!!!')
    

    class FirstPayment:
        def __init__(self, start: str, end: str=None):
            self.col = [
                '매출일',
                '주문번호',
                'PRODUCT_ID',
                'POLICY_ID'
            ]
            self.data = self.filtered_table(start, end)
            self.result = self.groupby()        
        
        def read(self):
            self.data = pd.read_parquet(DATA_PATH['toss_payment'])
            return self
        
        def filter_data(self, end: str=None):
            if end:
                condition = (self.data['매출일'] < end)
                self.data = self.data[condition]
            return self
        
        def add_month(self):
            self.data['month'] = pd.to_datetime(self.data['매출일']).dt.to_period('M')
            return self
        
        def first_payment(self):
            min = self.data.groupby(['POLICY_ID', 'PRODUCT_ID'])['month'].min().reset_index()
            min.rename(columns={'month': 'first_payment'}, inplace=True)
            return min
        
        def last_payment(self):
            max = self.data.groupby(['POLICY_ID', 'PRODUCT_ID'])['month'].max().reset_index()
            max.rename(columns={'month': 'last_payment'}, inplace=True)
            return max
        
        def merge(self):
            min = self.first_payment()
            max = self.last_payment()
            return pd.merge(min, max, on=['POLICY_ID', 'PRODUCT_ID'], how='left')
        
        def filter_merge(self, start: str):
            result = self.merge()
            return result[(result['first_payment'] == pd.to_datetime(start).to_period('M'))]
        
        def filtered_table(self, start, end):
            self.read().filter_data(end).add_month()
            return self.filter_merge(start)
        
        def groupby(self):
            group = self.data.groupby('PRODUCT_ID')['POLICY_ID'].count().reset_index()
            group.rename(columns={'POLICY_ID': '최초결제건수'}, inplace=True)
            return group
    
    
    class ExpiredProduct:
        '''
        상품별 만기 수량 집계
        
        Parameter
        ---
        date : '2023-08-01' 형식 / 마감 월에 맞춰 입력
        '''
        member_table_path = DATA_PATH['member_list']
        info_table_path = DATA_PATH['program_info']
        
        def __init__(self, start: str, end: str):
            print('만기 상품 집계 시작')
            self.expiration_group = self.groupby_expiration(start, end)
        
        def date_transform(self):
            date = datetime.strptime(self.date, '%Y-%m-%d')
            month = datetime.strftime(date, '%Y-%m')
            return month
        
        def __basic_data(self):
            print('회원 정보 추출 중...')
            col = ['상품정보', 'POLICY_ID', '보험가입일', '프로모션유무']
            df = pd.read_parquet(self.member_table_path)[col]
            df = df[(df['프로모션유무'] == 'Y')]
            
            print('프로그램 정보 추출 중...')
            col_ = ['PROGRAM_CODE', 'PROMOTION_MONTH']
            info = pd.read_parquet(self.info_table_path)[col_]
            info.rename(columns={'PROGRAM_CODE': '상품정보'}, inplace=True)
            
            print('회원정보와 프로그램 정보를 병합 중...')
            mg = pd.merge(df, info, on='상품정보', how='left')
            mg.drop(columns='프로모션유무', inplace=True)
            return mg
        
        def __add_expiration_date(self, start, end):
            '''
            상품만기 정보 컬럼 추가
            '''
            mg = self.__basic_data()

            # 만기일
            print('만기월 계산 중...')
            mg['보험가입일'] = pd.to_datetime(mg['보험가입일'])
                        
            selection = [
                mg['PROMOTION_MONTH'] == 36,
                mg['PROMOTION_MONTH'] == 24,
                mg['PROMOTION_MONTH'] == 12,
                mg['PROMOTION_MONTH'] == 6,
                mg['PROMOTION_MONTH'] == 3
            ]

            values = [
                mg['보험가입일'] + pd.DateOffset(months=36),
                mg['보험가입일'] + pd.DateOffset(months=24),
                mg['보험가입일'] + pd.DateOffset(months=12),
                mg['보험가입일'] + pd.DateOffset(months=6),
                mg['보험가입일'] + pd.DateOffset(months=3)
            ]

            mg['보험만기일'] = np.select(selection, values, mg['보험가입일'])
            
            # 세부 날짜 지정(현재 기준 만기 도래하지 않은 항목 삭제)
            mg = mg[mg['보험만기일'] < end]
            mg['보험만기일'] = mg['보험만기일'].dt.to_period('M')
            mg = mg[mg['보험만기일'] == pd.to_datetime(start).to_period('M')]
            mg['보험만기일'] = mg['보험만기일'].astype('str')
            return mg
        
        def groupby_expiration(self, start, end):
            mg = self.__add_expiration_date(start, end)
            print('만기월 그룹화 중...')
            mg = mg.groupby('상품정보')['보험만기일'].count().reset_index()
            return mg