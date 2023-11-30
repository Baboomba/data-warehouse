from config import DATA_PATH
from data.columns import case_transition, untact_column
from datetime import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd



class PaidProductCount:
    '''
    유상 마감 연산 클래스
    
    PARAMETERS
    ---
    save : 병합 파일 엑셀 저장 여부
    '''
    def __init__(self, save: bool):
        self.dataframe = self.return_merged_table()
        self.save_dataframe(save)
        # self.pivot_dic = self.create_result(self.dataframe)  # 엑셀 피벗으로 대체
        # self.save_files(save)   # 엑셀 피벗으로 대체 self.create_result() 위한 저장 함수
    
    class BasicData:
        def __init__(self) -> None:
            self.data = self.read_data()
        
        def read_data(self):
            _ss_col = [
                'PAYMENT_ID',
                'PACK_CODE',
                'INSURANCE_COMPANY_CODE',
                'PRODUCT_ID',
                'POLICY_ID'
            ]
            
            _ts_col = [
                '상점아이디(MID)',
                '정산액 입금일',
                '매출일',
                '결제일',
                '주문번호',
                '결제상태',
                '결제·취소액 (A)',
                'PG수수료 (B)',
                '당일 정산액 (C) = (A-B)'
            ]
            
            _cate_col = [
                '상품정보',
                '제품군_2',
                '제품시리즈',
                '제품시리즈_2',
                '보장타입',
                '유무상'
            ]
            
            ts = pd.read_parquet(DATA_PATH['toss_raw'])[_ts_col]
            ss = pd.read_parquet(DATA_PATH['samsung_raw'])[_ss_col]
            cate = pd.read_parquet(DATA_PATH['program_category'])[_cate_col]
            return [ts, ss, cate]
        
    class Merge(BasicData):
        def __init__(self):
            super().__init__()
            self.merge_process()
        
        def samsung(self):
            self.dataframe = pd.merge(
                left=self.data[0],
                right=self.data[1],
                left_on='주문번호',
                right_on='PAYMENT_ID',
                how='left'
            )
            return self
        
        def adjust(self):
            self.dataframe.drop_duplicates(inplace=True)
            self.dataframe.reset_index(drop=True, inplace=True)
            return self
        
        def check_count(self):
            if len(self.dataframe) != len(self.data[0]):
                raise Exception('The number of dataframe is different after merging')
            return self
        
        def category(self):
            self.dataframe = pd.merge(
                left=self.dataframe,
                right=self.data[2],
                left_on='PRODUCT_ID',
                right_on='상품정보',
                how='left'
            )
            return self
        
        def merge_process(self):
            self.samsung().adjust().check_count().category().adjust().check_count()
                    
    class ProcessAdditionalData(Merge):
        def __init__(self, save: bool=False):
            super().__init__()
            self.result(save)
        
        def alter_packcode(self):
            code = pd.DataFrame(case_transition)
            size = len(code)

            for num in range(size):
                con1 = (self.dataframe['PACK_CODE'] == code['code_before'][num])
                con2 = (self.dataframe['결제·취소액 (A)'] != code['price_before'][num])
                condition = (con1 & con2)
                self.dataframe.loc[condition, 'PACK_CODE'] = code['code_after'][num]        

            return self
        
        def process_offset(self):
            '''
            결제/취소 쌍 상계처리
            '''
            print('병합 데이터 결제/취소 상계 건 검색...')
            
            con1 = self.dataframe['주문번호'].duplicated(keep=False)
            con2 = self.dataframe['PAYMENT_ID'].isna()
            con3 = (self.dataframe['결제상태'].nunique() == 2)
            
            dupl = self.dataframe[con1 & con2 & con3].sort_values('주문번호')
            dupl.reset_index(inplace=True)

            _index = []

            for num in range(len(dupl) - 1):
                if dupl.loc[num, '주문번호'] != dupl.loc[num + 1, '주문번호']:
                    pass
                else:
                    if (dupl.loc[num, '결제·취소액 (A)'] != 0) and (dupl.loc[num + 1, '결제·취소액 (A)'] != 0):
                        if (dupl.loc[num, '결제·취소액 (A)'] + dupl.loc[num + 1, '결제·취소액 (A)']) == 0:
                            _index.append(dupl.loc[num, 'index'])
                            _index.append(dupl.loc[num + 1, 'index'])
                        
            print(f'상계 건 : {len(_index)}건')

            for idx in _index:
                self.dataframe.loc[idx, 'PACK_CODE'] = '`+-완료'
                print(f'인덱스 {idx}번 상계 처리')
                
            return self
        
        def previous_payment(self):
            '''
            전월 토스 결제건 처리
            '''
            pre = pd.read_parquet(DATA_PATH['toss_payment_merged'])
            pre = pre[['주문번호', 'packcode', 'ins', 'PRODUCT_ID']]
            pre.rename(columns={
                'packcode': 'PACK_CODE_',
                'ins': 'INSURANCE_COMPANY_CODE_',
                'PRODUCT_ID': 'PRODUCT_ID_'
            }, inplace=True)
            self.dataframe = pd.merge(self.dataframe, pre, how='left', on='주문번호')
            condition = ((self.dataframe['PACK_CODE'].isna()) & ~(self.dataframe['PACK_CODE_'].isna()))
            self.dataframe.loc[self.dataframe[condition].index, 'PACK_CODE'] = self.dataframe.loc[self.dataframe[condition].index, 'PACK_CODE_']
            self.dataframe.loc[self.dataframe[condition].index, 'INSURANCE_COMPANY_CODE'] = self.dataframe.loc[self.dataframe[condition].index, 'INSURANCE_COMPANY_CODE_']
            self.dataframe.loc[self.dataframe[condition].index, 'PRODUCT_ID'] = self.dataframe.loc[self.dataframe[condition].index, 'PRODUCT_ID_']
            self.dataframe.drop(columns=['PACK_CODE_', 'INSURANCE_COMPANY_CODE_', 'PRODUCT_ID_'], inplace=True)
            print('전월 토스 결제 처리 완료')
            return self
        
        def trim(self):
            self.dataframe.drop(columns='PAYMENT_ID', inplace=True)
            self.dataframe.drop(columns='상품정보', inplace=True)
            self.dataframe.drop_duplicates(inplace=True)
            return self
        
        def save(self, save: bool):
            if save:
                now = datetime.now().strftime('%Y%m%d')
                self.dataframe.to_excel(fr'result\유상마감데이터_{now}.xlsx')
        
        def result(self, save):
            self.alter_packcode().process_offset().previous_payment().trim().save(save)
    
    ## 여기부터 cell_style 메서드까지 미사용 코드    
    def create_result(self, df_join):
        '''
        보험사별 그룹화 집계 및 피벗
        
        상품정보와 pac_code 피벗테이블 딕셔너리 반환
        '''
        index = ['PRODUCT_ID', 'PACK_CODE']
        columns = ['결제상태', 'INSURANCE_COMPANY_CODE']
        values = '주문번호'
        
        groups = []
        print('그룹화 집계 중...')
        
        for element in index:
            print(f'그룹화 대상 : {element}, INSURANCE_COMPANY_CODE, 결제상태')
            col_pid = [element, 'INSURANCE_COMPANY_CODE', '결제상태']
            group = df_join.groupby(col_pid)['주문번호'].count().reset_index()
            groups.append(group)
            print(f'{element} 테이블 그룹화 집계 완료')
        
        pivots = {}
        print(f'피벗 테이블 생성...')

        for num, element in enumerate(index):
            pivot = groups[num].pivot(index=element, columns=columns, values=values).fillna(0)
            print(f'{element} 피벗 테이블 생성 완료')
        
            # 합계 컬럼 추가
            pivot[('합계', 'HW')] = pivot[('완료', 'HW')] - pivot[('취소', 'HW')]
            pivot[('합계', 'KB')] = pivot[('완료', 'KB')] - pivot[('취소', 'KB')]
            pivot[('합계', 'SS')] = pivot[('완료', 'SS')] - pivot[('취소', 'SS')]
            print(f'{element} 피벗 테이블 합계 컬럼 추가')
            
            pivots[element] = pivot
            
        return pivots
    
    def save_files(self, save):
        '''
        추출된 데이터를 엑셀 파일로 저장
        
        3개 시트 : 피벗테이블(pid, pac_code), 결제 테이블
        '''
        if save:
            wb = Workbook()
            print('엑셀 워크북 생성')
            self.pivot_dic['raw_data'] = self.dataframe
            
            for key, value in self.pivot_dic.items():
                sheet = wb.create_sheet(title=key)
                print(f'{key} 시트 생성')
                
                print('데이터 입력 중 : [', end='')
                ratio = value.shape[0] * 0.01
                for num, row in enumerate(dataframe_to_rows(value, index=True, header=True)):
                    sheet.append(row)
                    
                    if num >= ratio:
                        print('#', end='')
                        ratio += value.shape[0] * 0.01
                print(f']\n{key} 시트 입력 완료')
            
            delete = wb['Sheet']
            wb.remove(delete)
            wb = self.cell_style(wb)
            
            print('데이터 저장 중...')
            now = datetime.now().strftime('%Y%m%d')
            dir = fr'result\유상데이터정산_{now}.xlsx'
            wb.save(dir)
            print('데이터 저장 완료')
        
    def cell_style(self, workbook):
        '''
        셀 병합 및 폰트 사이즈 조정
        '''
        sheets = workbook.sheetnames
        
        # 범위 변수
        ranges = [
            'A1:A2',
            'B1:D1',
            'E1:G1',
            'H1:J1'
        ]
        
        for sht in sheets:
            # 시트 활성화
            ws = workbook[sht]
            workbook.active = ws
            print(f'{sht} 시트의 셀과 폰트를 조정 중...')
            
            if sht == 'raw_data':
                continue
            else:
                ws['A1'] = sht
                # 셀 병합
                for rng in ranges:
                    ws.merge_cells(rng)

            # 모든 데이터를 한 칸씩 올리기
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.value = ws.cell(row=cell.row + 1, column=cell.column).value

            # 폰트 조정
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(size=9)
            
        return workbook


#####################################################################################################################
## 무상 마감
#####################################################################################################################


class ConversionClassification:
    '''
    보험금 유상 전환 구분
    
    인스턴스 생성 시 회원넘버, 접수번호 데이터프레임 입력
    '''
    def __init__(self, dataframe, save: bool):
        self.basic_data = dataframe
        self.member_path = DATA_PATH['member_close']
        self.save_data(save)
        
    def read_member(self):
        return pd.read_parquet(self.member_path)
    
    def adjust_columns(self):
        df = self.read_member()
        return df[['POLICY_ID', '유무상 구분', '보험가입일', '무상 종료일']]
    
    def alter_type(self):
        df = self.adjust_columns()
        df['무상 종료일'] = pd.to_datetime(df['무상 종료일'])
        return df
    
    # processing samsung data
    def extract_date(self):
        self.basic_data['접수날짜'] = self.basic_data['접수번호'].str[:8]
        self.basic_data['접수날짜'] = pd.to_datetime(self.basic_data['접수날짜'], format='%Y%m%d')
        self.basic_data.rename(columns={'회원 넘버':'POLICY_ID'}, inplace=True)
        return self
    
    def merge_dataframe(self):
        self.extract_date()
        df = self.alter_type()
        return pd.merge(self.basic_data, df, on='POLICY_ID', how='left')
    
    def classify_payment(self):
        con = self.merge_dataframe()
        con['무상 종료일'] = pd.to_datetime(np.where(con['유무상 구분'] == '유상', None, con['무상 종료일']))
        return con
    
    def classify_timing(self):
        con = self.classify_payment()
        con['무상종료 전/후'] = (con['접수날짜'] - con['무상 종료일']).dt.days
        # condition to classify the time of the benefit
        conditions = [
            con['무상종료 전/후'] > 0,
            con['무상종료 전/후'] <= 0
        ]
        values = ['후', '전']
        
        con['무상종료 전/후'] = np.select(conditions, values, None)
        return con
    
    def correct_conversion(self):
        con = self.classify_timing()
        condition = ((con['유무상 구분'] == '전환') & (con['무상종료 전/후'] == '전'))
        con['유무상 구분'] = np.where(condition, '무상', con['유무상 구분'])
        return con
    
    def save_data(self, save: bool):
        if save:
            df = self.correct_conversion()
            df.to_excel(r'result\보험금전환구분.xlsx')
    
    
class PromotionProductCount:
    def __init__(self, save: bool=False):
        self.dataframe = None
        self.category = None
        self.result(save)
    
    def filter_promotion(self):
        self.dataframe = pd.read_parquet(DATA_PATH['member_close'])
        return self
    
    def read_category(self):
        return self
    
    def merge(self):
        self.dataframe = pd.merge(
            self.dataframe,
            self.category,
            on='상품정보',
            how='left'
        )
        return self
    
    def groupby(self):
        group = [
            '보험사',
            '무상제품',
            '정산년도',
            '보험상태'
        ]
        gr = self.dataframe.groupby(group)['POLICY_ID'].count()
        self.dataframe = pd.DataFrame(gr).reset_index(drop=True)
        self.dataframe.rename(columns={'POLICY_ID':'수량'}, inplace=True)
        return self
    
    def save(self, save: bool):
        if bool:
            pass
        pass
    
    def result(self, save):
        self.read_category().merge().groupby().save(save)


class PromotionClose:
    '''
    보험사 증빙용 무상상품 정산 마감
    '''
    def __init__(self):
        self.member_close_table = DATA_PATH['member_close']
        self.promotion_info_table = DATA_PATH['promotion_info']
        self.columns = [
            '상품정보',
            '프로그램명',
            '월요금',
            'POLICY_ID',
            '보험가입일',
            '보험해지일',
            '보험사',
            '보험상태',
            '유무상 구분'
        ]
        self.data = pd.read_parquet(self.member_close_table)
        self.final_result()
    
    def date_type_transform(self):
        self.data['보험가입일'] = pd.to_datetime(self.data['보험가입일'], format='%Y-%m-%d')
        self.data['보험해지일'] = pd.to_datetime(self.data['보험해지일'], format='%Y-%m-%d')
        return self
    
    def select_promotion(self):
        self.data = self.data[self.data['유무상 구분'] == '무상']
        return self
    
    def select_active(self):
        self.data = self.data[self.data['보험상태'] != '해지']
        return self
    
    def add_close_month(self):
        '''
        정산월 조정(하루씩 더함)
        
        2023년 9월 조정은 특이 사항임
        '''
        self.data['정산월'] = (self.data['보험가입일'] + pd.DateOffset(days=1)).dt.to_period('M')
        cond1 = (self.data['보험가입일'] > '2023-09-26')   # 9월 마감이 26일 일어남
        cond2 = (self.data['보험가입일'] < '2023-10-31')
        self.data.loc[cond1 & cond2, '정산월'] = pd.to_datetime('2023-10').strftime('%Y-%m')
        return self
    
    def group_for_result(self):
        self.data = self.data.groupby(['정산월', '상품정보', '보험사'])['POLICY_ID'].count().reset_index()
        return self
    
    def extract_promotion_info(self):
        info = pd.read_parquet(self.promotion_info_table)
        info = info[['상품정보', '프로모션명_요약']]
        info = info[~info['상품정보'].duplicated()]
        return info
    
    def final_result(self):
        self.date_type_transform().select_promotion().select_active().add_close_month().group_for_result()
        promotion = self.extract_promotion_info()
        self.data = pd.merge(self.data, promotion, on='상품정보', how='left')
        self.data.rename(columns={'POLICY_ID':'수량', '상품정보':'프로그램명'}, inplace=True)
        now = datetime.now().strftime('%Y%m%d')
        self.data.to_excel(fr'result\무상상품마감_{now}.xlsx')
        return self




#####################################################################################################################
## 비대면솔루션
#####################################################################################################################

class UntactSolution:
    '''
    Parameter
    ---
    period_dic : 정산 기간 딕셔너리.
    >>> date_dic = {'2023-10':['2023-09-26', '2023-10-30']}
    save : 결과 엑셀 저장 / default = False
    
    Atrribute
    ---
    result : 결과 데이터프레임 | start : 정산 시작 기준일 | end : 정산 마감 기준일
    '''
    def __init__(self, period_dic: dict, save: bool=False):
        self.result = self.assebled_function(period_dic)
        self.save(self.result, save)
    
    def assebled_function(self, period_dic):
        self.set_date(period_dic)
        solution = UntactSolution.JoinTime(self.start, self.end)
        df = solution.dataframe
        period = UntactSolution.PeriodAdjustment(df, period_dic)
        df = period.dataframe
        df = self.add_target_column(df)
        df = self.select_columns(df)
        df = self.classify_payment(df)
        return self.untactsolution_only(df)
        
    
    def save(self, df, save: bool) -> None:
        if save:
            now = datetime.now().strftime('%Y%m%d')
            df.to_excel(fr'result\유상 상품 비대면진단_{self.start}_{self.end}_{now}.xlsx')
            print('데이터 저장 완료')
        
    
    def set_date(self, period_dic: dict):
        i = 0
        for _, item in period_dic.items():
            if i == 0:
                self.start = item[0]
            if i == (len(period_dic) - 1):
                self.end = item[1]
            i += 1
    
    class JoinTime:
        '''
        가입 시점 계산
        
        보험가입일 - 최초통화일 <= 2  --> 3일 이내
        
        보험가입일 - 최초통화일 >= 3  --> 최초통화일+2일 이후
        '''
        def __init__(self, start: str, end: str):
            self.cols = ['최초통화일', '보험가입일', '보험해지일']
            self.dataframe = self.read_data()
            self.produce_result(start, end)
        
        def read_data(self):
            return pd.read_parquet(DATA_PATH['member_list'])
        
        def alter_type(self):
            for col in self.cols:
                self.dataframe[col] = pd.to_datetime(self.dataframe[col], format='%Y-%m-%d')
            return self
        
        def select_data(self, start: str, end: str):
            cond_1 = (self.dataframe['보험가입일'] >= start)
            cond_2 = (self.dataframe['보험가입일'] <= end)
            self.dataframe = self.dataframe[cond_1 & cond_2]
            return self
        
        def add_timing(self):
            self.dataframe['가입 시점'] = (self.dataframe['보험가입일'] - self.dataframe['최초통화일']).dt.days
            return self
        
        def categorize_timing(self):
            condition = (self.dataframe['가입 시점'] <= 2)
            self.dataframe['가입 시점'] = np.where(condition, '3일 이내', '최초통화일+2일 이후')
            return self
        
        def adjust_dataframe(self):
            self.dataframe.reset_index(drop=True, inplace=True)
            self.dataframe.drop('최초통화일', axis=1, inplace=True)
            return self
        
        def produce_result(self, start, end):
            self.alter_type().select_data(start, end).add_timing().categorize_timing().adjust_dataframe()
    
    class PeriodAdjustment:
        def __init__(self, dataframe, period_dic: dict):
            self.cols = {'보험가입일':'가입월', '보험해지일':'해지월'}
            self.dataframe = dataframe
            self.produce_result(period_dic)
        
        def add_columns(self):
            self.dataframe[['가입월', '해지월']] = None
            return self
        
        def adjust_period(self, period_dic: dict):
            for key, item in period_dic.items():
                for day, month in self.cols.items():
                    cond_start = (self.dataframe[day] >= item[0])
                    cond_end = (self.dataframe[day] <= item[1])
                    self.dataframe.loc[self.dataframe[(cond_start & cond_end)].index, month] = key
            return self
                
        def alter_type(self):
            for col in self.cols:
                self.dataframe[col] = self.dataframe[col].dt.strftime('%Y-%m-%d')
            return self
        
        def produce_result(self, period_dic):
            self.add_columns().adjust_period(period_dic).alter_type()
        
        
    def add_target_column(self, df):
        print('정산 대상 여부 계산 중...')
        cond1 = (df['가입월'] == df['해지월'])
        df['정산대상여부'] = np.where(cond1, '당월해지', '정산대상')
        return df
    
    def select_columns(self, df):
        print('컬럼 조정 중...')
        df = df.rename(columns=untact_column)
        col = list(untact_column.values())
        df = df[col]
        df = df.set_index('PROGRAM_CODE', drop=True)
        return df
    
    def classify_payment(self, df):
        condition = (df['상품 구분'] == 'Y')
        df['상품 구분'] = np.where(condition, '무상', '유상')
        return df[df['상품 구분'] == '유상']
    
    def untactsolution_only(self, df):
        return df[df['비대면솔루션 값'].notnull()]